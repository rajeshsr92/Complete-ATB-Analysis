import sys
import os

def _find_and_add_lib():
    candidates = [
        os.path.join(os.path.expanduser('~'), 'AppData', 'Local', 'medicare_dash_lib'),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'lib'),
        os.environ.get('MEDICARE_LIB_PATH', ''),
    ]
    for path in candidates:
        path = os.path.normpath(path)
        if os.path.isdir(path) and path not in sys.path:
            if os.path.exists(os.path.join(path, 'flask', '__init__.py')):
                sys.path.insert(0, path)
                return path
    return None

_find_and_add_lib()

try:
    import pyarrow as _pa
    _STUBS = ['__version__', 'Array', 'ChunkedArray', 'Table', 'RecordBatch', 'Schema', 'Field', 'DataType']
    if not hasattr(_pa, '__version__'):
        _pa.__version__ = '0.0.0'
    for _a in _STUBS[1:]:
        if not hasattr(_pa, _a):
            setattr(_pa, _a, type(_a, (), {}))
except ImportError:
    pass

import io
import datetime
import threading
import pandas as pd
from flask import Flask, render_template, jsonify, request, send_file
from data_loader import (load_all_atb_files, discover_clients, get_filter_values,
                          get_billing_entities, load_production_file, load_work_queue_file)
from analytics import (wow_trending, trending_summary,
                        atb_retention_analysis,
                        aging_migration, rollover_summary, migration_cell_detail,
                        atb_bifurcation, bifurcation_summary,
                        unbilled_analysis, balance_group_breakdown, aging_velocity,
                        aging_contributors, compute_high_dollar_threshold,
                        denial_analysis, denial_velocity,
                        cash_collection_action_plan,
                        get_priority_encounter_df,
                        untouched_claims_analysis,
                        OVER_90_BUCKETS, BUCKET_INDEX, _DENIAL_EMPTY)

if getattr(sys, 'frozen', False):
    _meipass = sys._MEIPASS
    app = Flask(
        __name__,
        template_folder=os.path.join(_meipass, 'dashboard', 'templates'),
        static_folder=os.path.join(_meipass, 'dashboard', 'static'),
    )
else:
    app = Flask(__name__)

@app.after_request
def no_cache(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    return response

_clients = {}
_clients_lock = threading.Lock()

_production_dfs  = {}
_production_lock = threading.Lock()

_work_queue_dfs  = {}
_work_queue_lock = threading.Lock()


def _get_production_df(client_name):
    with _production_lock:
        if client_name not in _production_dfs:
            _production_dfs[client_name] = load_production_file(client_name)
        return _production_dfs[client_name]


def _get_work_queue_df(client_name):
    with _work_queue_lock:
        if client_name not in _work_queue_dfs:
            _work_queue_dfs[client_name] = load_work_queue_file(client_name)
        return _work_queue_dfs[client_name]


def _inject_pct(points, total_ar):
    """Add pct = balance / total_ar * 100 to each point that carries a 'balance' key."""
    if not total_ar:
        return
    for p in points:
        if 'balance' in p and p['balance'] is not None and 'pct' not in p:
            p['pct'] = round(float(p['balance']) / float(total_ar) * 100, 1)


def _get_state(name):
    with _clients_lock:
        return _clients.get(name)


def _load_client(client_name, atb_folder):
    with _clients_lock:
        if client_name not in _clients:
            _clients[client_name] = {
                'weekly_data': {}, 'weeks': [], 'loading': True,
                'load_log': [], 'error': None
            }
    state = _clients[client_name]

    def log(msg):
        print(f'[{client_name}] {msg}')
        state['load_log'].append(msg)

    try:
        log('Starting data load...')
        if not isinstance(atb_folder, dict) and not os.path.isdir(atb_folder):
            raise FileNotFoundError(f'ATB folder not found: {atb_folder}')
        data = load_all_atb_files(atb_folder, progress_cb=log)
        state['weekly_data'] = data
        state['weeks'] = sorted(data.keys())
        log(f'Loaded {len(data)} weeks: {state["weeks"]}')
    except Exception as e:
        state['error'] = str(e)
        log(f'ERROR: {e}')
    finally:
        state['loading'] = False


def _load_all_clients():
    clients = discover_clients()
    if not clients:
        print('WARNING: No client folders found under Data/')
        return
    for c in clients:
        t = threading.Thread(target=_load_client,
                             args=(c['name'], c['atb_folder']), daemon=True)
        t.start()


def _reload_client(client_name, atb_folder):
    """Force-replaces client state and reloads all files. Safe to call while old load is running."""
    new_state = {'weekly_data': {}, 'weeks': [], 'loading': True, 'load_log': [], 'error': None}
    with _clients_lock:
        _clients[client_name] = new_state
    state = new_state

    def log(msg):
        print(f'[{client_name}] {msg}')
        state['load_log'].append(msg)

    try:
        log('Starting data reload...')
        if not isinstance(atb_folder, dict) and not os.path.isdir(atb_folder):
            raise FileNotFoundError(f'ATB folder not found: {atb_folder}')
        data = load_all_atb_files(atb_folder, progress_cb=log)
        state['weekly_data'] = data
        state['weeks'] = sorted(data.keys())
        log(f'Loaded {len(data)} weeks: {state["weeks"]}')
    except Exception as e:
        state['error'] = str(e)
        log(f'ERROR: {e}')
    finally:
        state['loading'] = False


def _resolve_client(name=None):
    """Return (client_name, state) or (None, error_tuple)."""
    if not name:
        with _clients_lock:
            for k, v in _clients.items():
                return k, v
        return None, (jsonify({'error': 'No clients loaded'}), 503)
    state = _get_state(name)
    if state is None:
        return None, (jsonify({'error': f'Client not found: {name}'}), 404)
    return name, state


def _apply_filters(df, req):
    """Apply RFC, Health Plan, Balance Type, Claim Status, DAC, and CTAC filters."""
    rfc  = [v for v in req.args.get('resp_fin_class', '').split(',') if v]
    rhp  = [v for v in req.args.get('resp_health_plan', '').split(',') if v]
    bt   = [v for v in req.args.get('balance_type', '').split(',') if v]
    cs   = [v for v in req.args.get('claim_status', '').split(',') if v]
    dac  = [v for v in req.args.get('dac', '').split(',') if v]
    ctac = [v for v in req.args.get('ctac', '').split(',') if v]
    if rfc:
        df = df[df['Responsible Financial Class'].astype(str).isin(rfc)]
    if rhp:
        df = df[df['Responsible Health Plan'].astype(str).isin(rhp)]
    if bt and 'Balance Type' in df.columns:
        df = df[df['Balance Type'].astype(str).str.strip().isin(bt)]
    if cs and 'Claim Status' in df.columns:
        df = df[df['Claim Status'].astype(str).isin(cs)]
    if dac and 'Discharge Aging Category' in df.columns:
        df = df[df['Discharge Aging Category'].astype(str).isin(dac)]
    if ctac and 'Claim Transmission Age Category' in df.columns:
        df = df[df['Claim Transmission Age Category'].astype(str).isin(ctac)]
    return df


def _apply_all_filters(df, req):
    """Apply fin/plan filters, then optional high dollar filter."""
    df = _apply_filters(df, req)
    if req.args.get('high_dollar') == 'true':
        try:
            pct = float(req.args.get('hd_pct', 0.60))
            threshold = compute_high_dollar_threshold(df, pct)['threshold']
            if threshold > 0:
                df = df[df['Balance Amount'] >= threshold]
        except Exception as e:
            print(f'[HD filter error] {e}')
    return df


# ── routes ───────────────────────────────────────────────────────

@app.route('/')
def index():
    on_railway = bool(os.environ.get('RAILWAY_ENVIRONMENT'))
    return render_template('index.html', show_footer=on_railway)


@app.route('/api/clients')
def api_clients():
    discovered = discover_clients()
    result = []
    for c in discovered:
        state = _get_state(c['name']) or {}
        result.append({
            'name': c['name'],
            'loading': state.get('loading', True),
            'error': state.get('error'),
            'weeks': state.get('weeks', []),
        })
    return jsonify(result)


@app.route('/api/status')
def api_status():
    client = request.args.get('client')
    _, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    return jsonify({
        'loading': state['loading'],
        'error': state['error'],
        'weeks_loaded': len(state['weeks']),
        'weeks': state['weeks'],
        'log': state['load_log'][-5:],
    })


@app.route('/api/reload', methods=['POST'])
def api_reload():
    client_param = request.args.get('client')
    discovered = {c['name']: c for c in discover_clients()}

    if client_param:
        if client_param not in discovered:
            return jsonify({'error': f'Client not found: {client_param}'}), 404
        targets = [discovered[client_param]]
    else:
        targets = list(discovered.values())

    reloaded, skipped = [], []
    for c in targets:
        name = c['name']
        with _clients_lock:
            existing = _clients.get(name)
            currently_loading = existing is not None and existing.get('loading', False)
        if currently_loading:
            skipped.append(name)
        else:
            t = threading.Thread(target=_reload_client, args=(name, c['atb_folder']), daemon=True)
            t.start()
            reloaded.append(name)

    return jsonify({'reloaded': reloaded, 'skipped': skipped})


@app.route('/api/weeks')
def api_weeks():
    client = request.args.get('client')
    _, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    return jsonify({'weeks': state['weeks']})


@app.route('/api/filters')
def api_filters():
    """Return unique Responsible Financial Class and Health Plan values for dropdowns."""
    client = request.args.get('client')
    name, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503
    return jsonify(get_filter_values(state['weekly_data']))


@app.route('/api/trending')
def api_trending():
    client = request.args.get('client')
    _, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503
    filtered = {w: _apply_all_filters(df, request) for w, df in state['weekly_data'].items()}
    rows = wow_trending(filtered)
    summary = trending_summary(rows)
    weeks = state['weeks']
    if weeks:
        total_ar = float(filtered[weeks[-1]]['Balance Amount'].sum())
        _inject_pct(summary, total_ar)
    return jsonify({'rows': rows, 'summary': summary})


@app.route('/api/migration')
def api_migration():
    client = request.args.get('client')
    _, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503
    from_week = request.args.get('from')
    to_week = request.args.get('to')
    if not from_week or not to_week:
        return jsonify({'error': 'from and to parameters required'}), 400
    wd = state['weekly_data']
    if from_week not in wd or to_week not in wd:
        return jsonify({'error': 'Week not found'}), 404
    a = _apply_all_filters(wd[from_week], request)
    b = _apply_all_filters(wd[to_week], request)
    result = aging_migration(a, b)
    summary_pts = rollover_summary(result)
    _inject_pct(summary_pts, float(b['Balance Amount'].sum()))
    result['summary_points'] = summary_pts
    return jsonify(result)


@app.route('/api/retention')
def api_retention():
    client = request.args.get('client')
    _, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503
    from_week = request.args.get('from')
    to_week   = request.args.get('to')
    if not from_week or not to_week:
        return jsonify({'error': 'from and to parameters required'}), 400
    wd = state['weekly_data']
    if from_week not in wd or to_week not in wd:
        return jsonify({'error': 'Week not found'}), 404
    a = _apply_all_filters(wd[from_week], request)
    b = _apply_all_filters(wd[to_week], request)
    return jsonify(atb_retention_analysis(a, b))


@app.route('/api/migration/detail')
def api_migration_detail():
    client = request.args.get('client')
    _, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503
    from_week   = request.args.get('from')
    to_week     = request.args.get('to')
    from_bucket = request.args.get('from_bucket') or None
    to_bucket   = request.args.get('to_bucket') or None
    if not from_week or not to_week:
        return jsonify({'error': 'from and to parameters required'}), 400
    wd = state['weekly_data']
    if from_week not in wd or to_week not in wd:
        return jsonify({'error': 'Week not found'}), 404
    a = _apply_all_filters(wd[from_week], request)
    b = _apply_all_filters(wd[to_week], request)
    rows = migration_cell_detail(a, b, from_bucket, to_bucket)
    return jsonify({'rows': rows, 'count': len(rows)})


@app.route('/api/bifurcation')
def api_bifurcation():
    client = request.args.get('client')
    _, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503
    week = request.args.get('week')
    if not week:
        return jsonify({'error': 'week parameter required'}), 400
    wd = state['weekly_data']
    weeks = state['weeks']
    if week not in wd:
        return jsonify({'error': 'Week not found'}), 404
    idx = weeks.index(week)
    prior_week = weeks[idx - 1] if idx > 0 else week
    curr = _apply_all_filters(wd[week], request)
    prior = _apply_all_filters(wd[prior_week], request)
    bifur_result = atb_bifurcation(curr, prior)
    unbilled = unbilled_analysis(curr, prior)
    bifur_result['unbilled'] = unbilled
    summary_pts = bifurcation_summary(bifur_result, unbilled)
    _inject_pct(summary_pts, float(curr['Balance Amount'].sum()))
    bifur_result['summary_points'] = summary_pts
    return jsonify(bifur_result)


@app.route('/api/aging-contributors')
def api_aging_contributors():
    client = request.args.get('client')
    _, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503
    week = request.args.get('week')
    wd = state['weekly_data']
    weeks = state['weeks']
    if not weeks:
        return jsonify({'error': 'No data'}), 404
    if not week or week not in wd:
        week = weeks[-1]
    idx = weeks.index(week)
    prior_week = weeks[idx - 1] if idx > 0 else week
    curr = _apply_all_filters(wd[week], request)
    prior = _apply_all_filters(wd[prior_week], request)
    top_n = int(request.args.get('top_n', 15))
    result = aging_contributors(curr, prior, top_n=top_n)
    _inject_pct(result.get('key_points', []), float(curr['Balance Amount'].sum()))
    return jsonify(result)


@app.route('/api/high-dollar-threshold')
def api_hd_threshold():
    client = request.args.get('client')
    _, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503
    weeks = state['weeks']
    if not weeks:
        return jsonify({'error': 'No data'}), 404
    week = request.args.get('week')
    if not week or week not in state['weekly_data']:
        week = weeks[-1]
    df = _apply_filters(state['weekly_data'][week], request)
    pct = float(request.args.get('hd_pct', 0.60))
    result = compute_high_dollar_threshold(df, pct)
    result['week'] = week
    return jsonify(result)


@app.route('/api/billing-entities')
def api_billing_entities():
    client = request.args.get('client')
    name, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503
    return jsonify({'entities': get_billing_entities(state['weekly_data'])})


@app.route('/api/unbilled')
def api_unbilled():
    client = request.args.get('client')
    _, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503
    week = request.args.get('week')
    wd = state['weekly_data']
    weeks = state['weeks']
    if not weeks:
        return jsonify({'error': 'No data'}), 404
    if not week or week not in wd:
        week = weeks[-1]
    idx = weeks.index(week)
    prior_week = weeks[idx - 1] if idx > 0 else week
    curr = _apply_all_filters(wd[week], request)
    prior = _apply_all_filters(wd[prior_week], request)
    return jsonify(unbilled_analysis(curr, prior))


@app.route('/api/balance-groups')
def api_balance_groups():
    client = request.args.get('client')
    _, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503
    week = request.args.get('week')
    wd = state['weekly_data']
    weeks = state['weeks']
    if not weeks:
        return jsonify({'error': 'No data'}), 404
    if not week or week not in wd:
        week = weeks[-1]
    idx = weeks.index(week)
    prior_week = weeks[idx - 1] if idx > 0 else week
    curr = _apply_all_filters(wd[week], request)
    prior = _apply_all_filters(wd[prior_week], request)
    return jsonify(balance_group_breakdown(curr, prior))


@app.route('/api/aging-velocity')
def api_aging_velocity():
    client = request.args.get('client')
    _, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503
    week = request.args.get('week')
    wd = state['weekly_data']
    weeks = state['weeks']
    if not weeks:
        return jsonify({'error': 'No data'}), 404
    if not week or week not in wd:
        week = weeks[-1]
    curr = _apply_all_filters(wd[week], request)
    return jsonify(aging_velocity(curr))


@app.route('/api/denials')
def api_denials():
    client = request.args.get('client')
    _, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503
    week = request.args.get('week')
    wd = state['weekly_data']
    weeks = state['weeks']
    if not weeks:
        return jsonify({'error': 'No data'}), 404
    if not week or week not in wd:
        week = weeks[-1]
    idx = weeks.index(week)
    prior_week = weeks[idx - 1] if idx > 0 else week
    curr  = _apply_all_filters(wd[week], request)
    prior = _apply_all_filters(wd[prior_week], request)
    result = denial_analysis(curr, prior)
    _inject_pct(result.get('summary_points', []), float(curr['Balance Amount'].sum()))
    return jsonify(result)


@app.route('/api/denial-velocity')
def api_denial_velocity():
    client = request.args.get('client')
    _, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503
    filtered = {w: _apply_all_filters(df, request) for w, df in state['weekly_data'].items()}
    result = denial_velocity(filtered)
    weeks = state['weeks']
    if weeks:
        _inject_pct(result.get('summary_points', []), float(filtered[weeks[-1]]['Balance Amount'].sum()))
    return jsonify(result)


@app.route('/api/cash-action-plan')
def api_cash_action_plan():
    client = request.args.get('client')
    _, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503
    week  = request.args.get('week')
    wd    = state['weekly_data']
    weeks = state['weeks']
    if not weeks:
        return jsonify({'error': 'No data'}), 404
    if not week or week not in wd:
        week = weeks[-1]
    idx        = weeks.index(week)
    prior_week = weeks[idx - 1] if idx > 0 else week
    filtered_all = {w: _apply_all_filters(df, request) for w, df in wd.items()}
    curr  = filtered_all[week]
    prior = filtered_all[prior_week]
    return jsonify(cash_collection_action_plan(filtered_all, curr, prior))


# ── download helpers ─────────────────────────────────────────────

def _make_excel_response(df, title, context_str, filename):
    """Return a Flask send_file response with a styled Excel workbook."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    ncols   = len(df.columns)
    last_col = get_column_letter(ncols) if ncols else 'A'

    # Row 1: dark-navy title bar
    ws.append([title] + [''] * max(ncols - 1, 0))
    ws.merge_cells(f'A1:{last_col}1')
    c = ws['A1']
    c.font      = Font(bold=True, color='FFFFFF', size=13)
    c.fill      = PatternFill('solid', fgColor='1E3A5F')
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 22

    # Row 2: context info
    ws.append([context_str] + [''] * max(ncols - 1, 0))
    ws.merge_cells(f'A2:{last_col}2')
    c = ws['A2']
    c.font      = Font(size=9, color='475569')
    c.fill      = PatternFill('solid', fgColor='F1F5F9')
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 15

    # Row 3: spacer
    ws.append([''])
    ws.row_dimensions[3].height = 5

    # Row 4: column headers
    ws.append(list(df.columns))
    for ci, col_name in enumerate(df.columns, 1):
        c = ws.cell(row=4, column=ci)
        c.font      = Font(bold=True, color='1E293B', size=10)
        c.fill      = PatternFill('solid', fgColor='E2E8F0')
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[4].height = 16

    # Sanitize: convert pd.NA / nullable-int NA → None so openpyxl can write them
    df = df.astype(object).where(df.notna(), other=None)

    # Data rows (row 5+)
    for ri, row_data in enumerate(df.itertuples(index=False), 5):
        ws.append(list(row_data))
        fill_color = 'FFFFFF' if ri % 2 == 1 else 'F8FAFC'
        for ci in range(1, ncols + 1):
            c = ws.cell(row=ri, column=ci)
            c.fill      = PatternFill('solid', fgColor=fill_color)
            c.font      = Font(size=9)
            c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[ri].height = 13

    # Auto-width (capped at 42) — df already sanitized above, so v is None or a real value
    for ci, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(ci)
        max_len = max(
            len(str(col_name)),
            max((len(str(v)) for v in df.iloc[:, ci - 1] if v is not None), default=0)
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 42)

    ws.freeze_panes = 'A5'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def _select_download_cols(df, priority_cols):
    """Re-order df so priority_cols come first; append remaining columns."""
    present = [c for c in priority_cols if c in df.columns]
    rest    = [c for c in df.columns if c not in set(present)]
    return df[present + rest]


def _download_context(client_name, week=None, extras=None):
    """Build a single-line context string for the Excel row 2."""
    parts = [f'Client: {client_name}']
    if week:
        parts.append(f'Week: {week}')
    rfc = [v for v in request.args.get('resp_fin_class', '').split(',') if v]
    rhp = [v for v in request.args.get('resp_health_plan', '').split(',') if v]
    if rfc:
        parts.append('Fin.Class: ' + ', '.join(rfc[:3]) + ('…' if len(rfc) > 3 else ''))
    if rhp:
        parts.append('Health Plan: ' + ', '.join(rhp[:3]) + ('…' if len(rhp) > 3 else ''))
    if request.args.get('high_dollar') == 'true':
        parts.append('High Dollar: ON')
    if extras:
        parts.extend(extras)
    parts.append(f'Generated: {datetime.date.today().isoformat()}')
    return '  |  '.join(parts)


# ── download routes ───────────────────────────────────────────────

@app.route('/api/download/trending')
def download_trending():
    client = request.args.get('client')
    name, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503
    wd    = state['weekly_data']
    weeks = state['weeks']
    if not weeks:
        return jsonify({'error': 'No data'}), 404
    week = request.args.get('week')
    if not week or week not in wd:
        week = weeks[-1]
    df = _apply_all_filters(wd[week], request)
    df = df.sort_values('Balance Amount', ascending=False).drop_duplicates('Encounter Number')
    priority = ['Encounter Number', 'Responsible Health Plan', 'Responsible Financial Class',
                'Balance Amount', 'Discharge Aging Category', 'Discharge Date',
                'Balance Group', 'Unbilled Aging Category']
    df  = _select_download_cols(df, priority)
    ctx = _download_context(name, week)
    return _make_excel_response(df, 'Volume Trending — Encounter Detail', ctx,
                                f'{name}_Trending_{week}.xlsx'.replace(' ', '_'))


@app.route('/api/download/migration')
def download_migration():
    client = request.args.get('client')
    name, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503
    from_week   = request.args.get('from')
    to_week     = request.args.get('to')
    from_bucket = request.args.get('from_bucket', '')
    to_bucket   = request.args.get('to_bucket', '')
    if not from_week or not to_week:
        return jsonify({'error': 'from and to parameters required'}), 400
    wd = state['weekly_data']
    if from_week not in wd or to_week not in wd:
        return jsonify({'error': 'Week not found'}), 404

    df_from = _apply_all_filters(wd[from_week], request)
    df_to   = _apply_all_filters(wd[to_week],   request)
    df_from = df_from.sort_values('Balance Amount', ascending=False).drop_duplicates('Encounter Number')
    df_to   = df_to.sort_values('Balance Amount', ascending=False).drop_duplicates('Encounter Number')

    from_encs = set(df_from['Encounter Number'].astype(str))
    to_encs   = set(df_to['Encounter Number'].astype(str))

    fname = f'{name}_Migration_{from_week}_to_{to_week}.xlsx'.replace(' ', '_').replace('/', '-')

    # ── Cell-specific download (single from→to bucket pair) ──────
    if from_bucket and to_bucket and from_bucket != 'NEW' and to_bucket != 'RESOLVED':
        merged = df_from.merge(
            df_to[['Encounter Number', 'Balance Amount', 'Discharge Aging Category']],
            on='Encounter Number', suffixes=('_from', '_to'), how='inner'
        ).rename(columns={
            'Balance Amount_from':           'Prior Week Balance',
            'Balance Amount_to':             'Current Week Balance',
            'Discharge Aging Category_from': 'Prior Week Bucket',
            'Discharge Aging Category_to':   'Current Week Bucket',
        })
        merged = merged[
            (merged['Prior Week Bucket'] == from_bucket) &
            (merged['Current Week Bucket'] == to_bucket)
        ].copy()
        merged['Balance Change']  = (merged['Current Week Balance'] - merged['Prior Week Balance']).round(2)
        merged['Migration Path']  = from_bucket + ' → ' + to_bucket
        fi = BUCKET_INDEX.get(from_bucket, 99)
        ti = BUCKET_INDEX.get(to_bucket, 99)
        if fi == ti:
            merged['Movement Type'] = 'Stayed (same bucket)'
        elif ti > fi:
            diff = ti - fi
            merged['Movement Type'] = f'Aged {diff} bucket{"s" if diff > 1 else ""} — WORSENED'
        else:
            merged['Movement Type'] = 'Improved (moved to younger bucket)'
        priority = ['Encounter Number', 'Migration Path', 'Movement Type',
                    'Responsible Health Plan', 'Responsible Financial Class',
                    'Prior Week Bucket', 'Current Week Bucket',
                    'Prior Week Balance', 'Current Week Balance', 'Balance Change', 'Discharge Date']
        result = _select_download_cols(merged, priority)
        title  = f'Aging Rollover — {from_bucket} → {to_bucket} Encounter Detail'
        ctx    = _download_context(name, to_week,
                                   [f'From: {from_week}', f'To: {to_week}',
                                    f'Migration: {from_bucket} → {to_bucket}',
                                    f'{len(result)} encounters'])
        return _make_excel_response(result, title, ctx, fname)

    # ── Full rollover report (all buckets with movement labels) ───
    # Section 1: continued encounters (inner join)
    to_cols = ['Encounter Number', 'Balance Amount', 'Discharge Aging Category']
    for c in ['Responsible Health Plan', 'Responsible Financial Class', 'Discharge Date']:
        if c in df_to.columns:
            to_cols.append(c)
    merged = df_from.merge(
        df_to[list(set(to_cols))],
        on='Encounter Number', suffixes=('_from', '_to'), how='inner'
    )
    # Rename columns
    rename_map = {
        'Balance Amount_from':           'Prior Week Balance',
        'Balance Amount_to':             'Current Week Balance',
        'Discharge Aging Category_from': 'Prior Week Bucket',
        'Discharge Aging Category_to':   'Current Week Bucket',
    }
    # Handle overlap when both DFs had the same col
    for col in ['Responsible Health Plan', 'Responsible Financial Class', 'Discharge Date']:
        if col + '_from' in merged.columns:
            rename_map[col + '_from'] = col
        if col + '_to' in merged.columns and col not in rename_map.values():
            rename_map[col + '_to'] = col
    merged = merged.rename(columns=rename_map)
    # Drop duplicate _to suffix columns
    merged = merged[[c for c in merged.columns if not c.endswith('_to')]]

    merged['Balance Change'] = (merged['Current Week Balance'] - merged['Prior Week Balance']).round(2)
    merged['Migration Path'] = merged['Prior Week Bucket'].astype(str) + ' → ' + merged['Current Week Bucket'].astype(str)

    def _movement_type(row):
        fi = BUCKET_INDEX.get(row['Prior Week Bucket'], 99)
        ti = BUCKET_INDEX.get(row['Current Week Bucket'], 99)
        if fi == ti:
            return 'Stayed (same bucket)'
        elif ti > fi:
            diff = ti - fi
            return f'Aged {diff} bucket{"s" if diff > 1 else ""} — WORSENED'
        else:
            return 'Improved (moved to younger bucket)'

    merged['Movement Type'] = merged.apply(_movement_type, axis=1)

    # Filter if only one bucket direction was specified
    if from_bucket:
        merged = merged[merged['Prior Week Bucket'] == from_bucket]
    if to_bucket:
        merged = merged[merged['Current Week Bucket'] == to_bucket]

    # Section 2: new encounters (in to_week but not from_week)
    new_df = df_to[~df_to['Encounter Number'].astype(str).isin(from_encs)].copy()
    new_df['Prior Week Balance']  = None
    new_df['Current Week Balance'] = new_df['Balance Amount']
    new_df['Balance Change']      = None
    new_df['Prior Week Bucket']   = 'N/A'
    new_df['Current Week Bucket'] = new_df['Discharge Aging Category']
    new_df['Migration Path']      = 'NEW → ' + new_df['Discharge Aging Category'].astype(str)
    new_df['Movement Type']       = 'New Encounter (not in prior week)'

    # Section 3: resolved encounters (in from_week but not to_week)
    res_df = df_from[~df_from['Encounter Number'].astype(str).isin(to_encs)].copy()
    res_df['Prior Week Balance']   = res_df['Balance Amount']
    res_df['Current Week Balance'] = None
    res_df['Balance Change']       = None
    res_df['Prior Week Bucket']    = res_df['Discharge Aging Category']
    res_df['Current Week Bucket']  = 'N/A'
    res_df['Migration Path']       = res_df['Discharge Aging Category'].astype(str) + ' → RESOLVED'
    res_df['Movement Type']        = 'Resolved / Removed from ATB'

    # Sort order for Movement Type
    _sort_key = {
        'Aged 1 bucket — WORSENED': 0,
        'Aged 2 buckets — WORSENED': 1,
        'Aged 3 buckets — WORSENED': 2,
        'Aged 4 buckets — WORSENED': 3,
        'Aged 5 buckets — WORSENED': 4,
        'Stayed (same bucket)': 10,
        'Improved (moved to younger bucket)': 20,
        'New Encounter (not in prior week)': 30,
        'Resolved / Removed from ATB': 40,
    }
    merged['_sort'] = merged['Movement Type'].map(lambda x: _sort_key.get(x, 5))
    merged = merged.sort_values(['_sort', 'Current Week Balance'], ascending=[True, False])
    merged = merged.drop(columns=['_sort'])

    all_parts = [merged]
    if not from_bucket and not to_bucket:
        new_df['_sort'] = 30
        res_df['_sort'] = 40
        new_df = new_df.sort_values('Current Week Balance', ascending=False).drop(columns=['_sort'], errors='ignore')
        res_df = res_df.sort_values('Prior Week Balance', ascending=False).drop(columns=['_sort'], errors='ignore')
        all_parts += [new_df, res_df]

    result = pd.concat(all_parts, ignore_index=True)

    priority = ['Encounter Number', 'Migration Path', 'Movement Type',
                'Responsible Health Plan', 'Responsible Financial Class',
                'Prior Week Bucket', 'Current Week Bucket',
                'Prior Week Balance', 'Current Week Balance', 'Balance Change', 'Discharge Date']
    result = _select_download_cols(result, priority)

    if from_bucket or to_bucket:
        label = f'{from_bucket or "All"} → {to_bucket or "All"}'
        title = f'Aging Rollover — {label} Encounter Detail'
        extras = [f'From: {from_week}', f'To: {to_week}', f'Migration: {label}']
    else:
        title  = f'Aging Migration — Full Rollover Report'
        extras = [f'From Week: {from_week}', f'To Week: {to_week}',
                  f'{len(merged)} continued  |  {len(new_df)} new  |  {len(res_df)} resolved']

    ctx = _download_context(name, to_week, extras)
    return _make_excel_response(result, title, ctx, fname)


@app.route('/api/download/bifurcation')
def download_bifurcation():
    client = request.args.get('client')
    name, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503
    wd    = state['weekly_data']
    weeks = state['weeks']
    if not weeks:
        return jsonify({'error': 'No data'}), 404
    week   = request.args.get('week')
    bucket = request.args.get('bucket', '')
    if not week or week not in wd:
        week = weeks[-1]
    df = _apply_all_filters(wd[week], request)
    df = df.sort_values('Balance Amount', ascending=False).drop_duplicates('Encounter Number')
    if bucket:
        df = df[df['Discharge Aging Category'] == bucket]
    priority = ['Encounter Number', 'Responsible Health Plan', 'Responsible Financial Class',
                'Balance Amount', 'Discharge Aging Category', 'Discharge Date', 'Unbilled Aging Category']
    df     = _select_download_cols(df, priority)
    label  = bucket or 'All Buckets'
    ctx    = _download_context(name, week, [f'Bucket: {label}'])
    title  = f'ATB Bifurcation — {label} Encounter Detail'
    fname  = f'{name}_Bifurcation_{week}_{label}.xlsx'.replace(' ', '_').replace('/', '-').replace('+', 'plus')
    return _make_excel_response(df, title, ctx, fname)


@app.route('/api/download/aging-contributors')
def download_aging_contributors():
    client = request.args.get('client')
    name, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503
    wd    = state['weekly_data']
    weeks = state['weeks']
    if not weeks:
        return jsonify({'error': 'No data'}), 404
    week       = request.args.get('week')
    health_plan = request.args.get('health_plan', '')
    fin_class   = request.args.get('fin_class', '')
    if not week or week not in wd:
        week = weeks[-1]
    df = _apply_all_filters(wd[week], request)
    df = df.sort_values('Balance Amount', ascending=False).drop_duplicates('Encounter Number')
    df = df[df['Discharge Aging Category'].isin(OVER_90_BUCKETS)]
    extras = ['Scope: 90+ Day Buckets']
    if health_plan:
        df = df[df['Responsible Health Plan'].astype(str) == health_plan]
        extras.append(f'Health Plan: {health_plan}')
    if fin_class:
        df = df[df['Responsible Financial Class'].astype(str) == fin_class]
        extras.append(f'Fin.Class: {fin_class}')
    priority = ['Encounter Number', 'Responsible Health Plan', 'Responsible Financial Class',
                'Balance Amount', 'Discharge Aging Category', 'Discharge Date']
    df    = _select_download_cols(df, priority)
    ctx   = _download_context(name, week, extras)
    fname = f'{name}_90plus_Contributors_{week}.xlsx'.replace(' ', '_')
    return _make_excel_response(df, '90+ Day Contributors — Encounter Detail', ctx, fname)


@app.route('/api/download/denials')
def download_denials():
    client = request.args.get('client')
    name, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503
    wd    = state['weekly_data']
    weeks = state['weeks']
    if not weeks:
        return jsonify({'error': 'No data'}), 404
    week        = request.args.get('week')
    denial_code = request.args.get('denial_code', '')
    if not week or week not in wd:
        week = weeks[-1]
    df = _apply_all_filters(wd[week], request)
    df = df.sort_values('Balance Amount', ascending=False).drop_duplicates('Encounter Number')
    _CODE = 'Last Denial Code and Reason'
    if _CODE not in df.columns:
        return jsonify({'error': 'No denial data available for this client/week'}), 404
    mask = df[_CODE].notna() & (~df[_CODE].astype(str).str.lower().str.strip().isin(_DENIAL_EMPTY))
    df   = df[mask].copy()
    extras = ['Scope: Open Denials Only']
    if denial_code:
        df = df[df[_CODE].astype(str) == denial_code]
        extras.append(f'Code: {denial_code[:60]}')
    _LDD = 'Last Denial Date'
    if _LDD in df.columns:
        ldd = pd.to_datetime(df[_LDD], errors='coerce')
        df['Days Since Last Denial'] = (pd.Timestamp.now() - ldd).dt.days.clip(lower=0).astype('Int64')
    priority = ['Encounter Number', 'Responsible Health Plan', 'Responsible Financial Class',
                'Balance Amount', 'Discharge Aging Category', _CODE,
                'Last Denial Group', _LDD, 'Days Since Last Denial', 'Discharge Date']
    df     = _select_download_cols(df, priority)
    ctx    = _download_context(name, week, extras)
    label  = (denial_code[:30] + '…') if denial_code and len(denial_code) > 30 else (denial_code or 'All Codes')
    title  = f'Open Denials — {label} Encounter Detail'
    fname  = f'{name}_Denials_{week}.xlsx'.replace(' ', '_')
    return _make_excel_response(df, title, ctx, fname)


@app.route('/api/download/denial-velocity')
def download_denial_velocity():
    client = request.args.get('client')
    name, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503
    wd    = state['weekly_data']
    weeks = state['weeks']
    if not weeks:
        return jsonify({'error': 'No data'}), 404
    week = request.args.get('week')
    if not week or week not in wd:
        week = weeks[-1]
    df = _apply_all_filters(wd[week], request)
    df = df.sort_values('Balance Amount', ascending=False).drop_duplicates('Encounter Number')
    _CODE = 'Last Denial Code and Reason'
    if _CODE not in df.columns:
        return jsonify({'error': 'No denial data available'}), 404
    mask = df[_CODE].notna() & (~df[_CODE].astype(str).str.lower().str.strip().isin(_DENIAL_EMPTY))
    df   = df[mask].copy()
    _LDD = 'Last Denial Date'
    if _LDD in df.columns:
        ldd = pd.to_datetime(df[_LDD], errors='coerce')
        df['Denial Age (Days)'] = (pd.Timestamp.now() - ldd).dt.days.clip(lower=0).astype('Int64')
    priority = ['Encounter Number', 'Responsible Health Plan', 'Responsible Financial Class',
                'Balance Amount', 'Discharge Aging Category', _CODE,
                'Last Denial Group', _LDD, 'Denial Age (Days)', 'Discharge Date']
    df    = _select_download_cols(df, priority)
    ctx   = _download_context(name, week, ['Scope: Denied Encounters'])
    title = f'Denial Velocity — Week {week} Encounter Detail'
    fname = f'{name}_DenialVelocity_{week}.xlsx'.replace(' ', '_')
    return _make_excel_response(df, title, ctx, fname)


_INSIGHT_FILTER_LABELS = {
    '90_denied':           '90+ Denied Claims',
    '61_90_rollover':      '61-90 Day Rollover Risk',
    'quick_win':           'Quick Win — Fresh Denials 91-120d',
    'timely_filing':       'Timely Filing Risk (CRITICAL & High)',
    'full_pool':           'Full Priority Pool',
    'dnfb':                'DNFB — Internal Billing',
    'aged_denials':        'Aged Denials 90+ Days',
    'top5_payers':         'Top 5 Payers by Priority',
    'ar_scope_highval':    'AR Scope High-Dollar (>$2K)',
    'claims_proc_highval': 'Claims Processing High-Dollar (>$2K)',
    'dnfb_client_highval': 'DNFB Client Scope High-Dollar (>$2K)',
}

# Prefix added to extras list for display; kept separate from ws.title (no colons allowed)
_INSIGHT_FILTER_SCOPE = {k: f'Scope: {v}' for k, v in _INSIGHT_FILTER_LABELS.items()}


def _get_claimed_encounters(df_curr, insight_filter):
    """Return encounter IDs already claimed by higher-priority insight downloads."""
    from analytics import INSIGHT_PRIORITY_ORDER
    fk_base = insight_filter.split(':')[0] if ':' in insight_filter else insight_filter
    if fk_base not in INSIGHT_PRIORITY_ORDER:
        return set()
    rank = INSIGHT_PRIORITY_ORDER.index(fk_base)
    claimed = set()
    for higher_fk in INSIGHT_PRIORITY_ORDER[:rank]:
        try:
            higher_df = get_priority_encounter_df(df_curr, higher_fk)
            claimed.update(higher_df['Encounter Number'].astype(str))
        except Exception:
            pass
    return claimed

@app.route('/api/download/cash-action-plan-all')
def download_cash_action_plan_all():
    """Download all Cash Collection Action Insights combined into one file."""
    client = request.args.get('client')
    name, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503
    wd    = state['weekly_data']
    weeks = state['weeks']
    if not weeks:
        return jsonify({'error': 'No data'}), 404
    week = request.args.get('week')
    if not week or week not in wd:
        week = weeks[-1]
    idx        = weeks.index(week)
    prior_week = weeks[idx - 1] if idx > 0 else week
    filtered_all = {w: _apply_all_filters(df, request) for w, df in wd.items()}
    curr  = filtered_all[week]
    prior = filtered_all[prior_week]

    plan     = cash_collection_action_plan(filtered_all, curr, prior)
    insights = plan.get('action_insights', [])
    if not insights:
        return jsonify({'error': 'No insights available'}), 404

    priority_cols = ['Insight Reason', 'Encounter Number', 'Responsible Health Plan',
                     'Responsible Financial Class', 'Balance Amount',
                     'Discharge Aging Category', 'Claim Status', 'Days Outstanding',
                     'Last Denial Code and Reason', 'Last Denial Date', 'Denial Age (Days)',
                     'TF Risk', 'Recommended Action', 'Discharge Date']

    seen_encs = set()
    chunks    = []
    for p in insights:
        fk      = p.get('filter_key', '') or ''
        # Use full insight text as the reason label so the user sees the complete description
        label   = p.get('text', '') or _INSIGHT_FILTER_LABELS.get(fk.split(':')[0] if ':' in fk else fk, fk)

        try:
            df_chunk = get_priority_encounter_df(curr, fk)
        except Exception:
            continue
        if df_chunk.empty:
            continue

        # Deduplicate: each encounter attributed to its first (highest-priority) insight
        mask = ~df_chunk['Encounter Number'].astype(str).isin(seen_encs)
        df_chunk = df_chunk[mask].copy()
        seen_encs.update(df_chunk['Encounter Number'].astype(str))
        if df_chunk.empty:
            continue

        df_chunk.insert(0, 'Insight Reason', label)
        chunks.append(df_chunk)

    if not chunks:
        return jsonify({'error': 'No encounter data'}), 404

    combined = pd.concat(chunks, ignore_index=True)
    # Sort: keep insight display order, then balance descending within each group
    reason_order = {(p.get('text', '') or ''): i for i, p in enumerate(insights)}
    combined['_sort'] = combined['Insight Reason'].map(reason_order).fillna(999)
    combined = combined.sort_values(['_sort', 'Balance Amount'], ascending=[True, False]).drop(columns=['_sort'])

    combined = _select_download_cols(combined, priority_cols)
    ctx   = _download_context(name, week, [f'All {len(insights)} Cash Collection Action Insights — Combined', 'Each encounter assigned to its highest-priority insight'])
    fname = f'{name}_CashActionInsights_All_{week}.xlsx'.replace(' ', '_')
    return _make_excel_response(combined, 'All Cash Collection Action Insights', ctx, fname)


@app.route('/api/download/cash-action-plan')
def download_cash_action_plan():
    client = request.args.get('client')
    name, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503
    wd    = state['weekly_data']
    weeks = state['weeks']
    if not weeks:
        return jsonify({'error': 'No data'}), 404
    week           = request.args.get('week')
    payer          = request.args.get('payer', '')
    insight_filter = request.args.get('insight_filter', '').strip()
    if not week or week not in wd:
        week = weeks[-1]
    curr = _apply_all_filters(wd[week], request)

    # insight_filter drives the data slice; legacy payer param still works
    active_filter = insight_filter or ('full_pool' if not payer else '')
    df = get_priority_encounter_df(curr, active_filter)

    extras = [_INSIGHT_FILTER_SCOPE.get(
        active_filter.split(':')[0] if ':' in active_filter else active_filter,
        'Scope: Priority Recovery Pool'
    )]
    if ':' in active_filter:
        extras.append(f'Filter: {active_filter.split(":", 1)[1]}')

    # top5_payers: filter pool to the named payers passed via filter_meta
    top5 = request.args.getlist('top5')
    if active_filter == 'top5_payers' and top5 and 'Responsible Health Plan' in df.columns:
        df = df[df['Responsible Health Plan'].astype(str).isin(top5)]

    # Legacy payer override (backwards-compatible)
    if payer and not insight_filter:
        df = df[df['Responsible Health Plan'].astype(str) == payer]
        extras.append(f'Payer: {payer}')

    # Deduplication: exclude encounters claimed by higher-priority insight downloads
    if insight_filter and active_filter not in ('full_pool', 'top5_payers'):
        excl = _get_claimed_encounters(curr, active_filter)
        if excl:
            before = len(df)
            df = df[~df['Encounter Number'].astype(str).isin(excl)]
            removed = before - len(df)
            if removed:
                extras.append(f'Deduped: {removed:,} enc. covered by higher-priority insights excluded')

    priority = ['Encounter Number', 'Responsible Health Plan', 'Responsible Financial Class',
                'Balance Amount', 'Discharge Aging Category', 'Claim Status', 'Days Outstanding',
                'Last Denial Code and Reason', 'Last Denial Date', 'Denial Age (Days)',
                'TF Risk', 'Recommended Action', 'Discharge Date']
    df    = _select_download_cols(df, priority)
    ctx   = _download_context(name, week, extras)
    fk_slug = (insight_filter or 'PriorityRecovery').replace(':', '_').replace(' ', '_')
    fname = f'{name}_{fk_slug}_{week}.xlsx'.replace(' ', '_')
    title = _INSIGHT_FILTER_LABELS.get(
        insight_filter.split(':')[0] if ':' in insight_filter else insight_filter,
        'Priority Recovery Score — Encounter Detail'
    )
    return _make_excel_response(df, title, ctx, fname)


_SNCA_CLIENT = 'Seneca_Health_SNCA_CA'


def _check_snca_client(name):
    """Return None if client is SNCA, otherwise a Flask error tuple."""
    if name and 'SNCA' in name.upper():
        return None
    return jsonify({'error': f'Production data is only available for {_SNCA_CLIENT}',
                    'snca_only': True}), 403


@app.route('/api/workables/untouched-claims')
def api_workables_untouched():
    client = request.args.get('client')
    name, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    err = _check_snca_client(name)
    if err:
        return err
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503

    prod_df = _get_production_df(name)
    if prod_df is None:
        return jsonify({'error': 'No production file found in Data/Production/'}), 404

    weeks = state['weeks']
    if not weeks:
        return jsonify({'error': 'No ATB data'}), 404
    week = request.args.get('week')
    if not week or week not in state['weekly_data']:
        week = weeks[-1]
    atb_df = _apply_all_filters(state['weekly_data'][week], request)

    date_str = request.args.get('date', '')
    try:
        end_date = pd.Timestamp(date_str) if date_str else pd.Timestamp.now().normalize()
    except Exception:
        end_date = pd.Timestamp.now().normalize()

    wq_df = _get_work_queue_df(name)
    exclude_wq = request.args.get('exclude_wq', 'false').lower() == 'true'

    try:
        result = untouched_claims_analysis(atb_df, prod_df, end_date, wq_df=wq_df, exclude_wq=exclude_wq)
    except Exception as e:
        print(f'[workables/untouched-claims ERROR] {e}')
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    try:
        _inject_pct(result['key_points'], result['summary']['total_bal'])
        result_slim = {k: v for k, v in result.items() if k != 'rows'}
        result_slim['rows']      = result['rows'][:500]
        result_slim['row_count'] = len(result['rows'])
        return jsonify(result_slim)
    except Exception as e:
        print(f'[workables/jsonify ERROR] {e}')
        import traceback; traceback.print_exc()
        return jsonify({'error': f'Serialization error: {e}'}), 500


@app.route('/api/download/workables-untouched')
def download_workables_untouched():
    client = request.args.get('client')
    name, state = _resolve_client(client)
    if isinstance(state, tuple):
        return state
    err = _check_snca_client(name)
    if err:
        return err
    if state['loading']:
        return jsonify({'error': 'Data still loading'}), 503

    prod_df = _get_production_df(name)
    if prod_df is None:
        return jsonify({'error': 'No production file found in Data/Production/'}), 404

    weeks = state['weeks']
    if not weeks:
        return jsonify({'error': 'No ATB data'}), 404
    week = request.args.get('week')
    if not week or week not in state['weekly_data']:
        week = weeks[-1]
    atb_df = _apply_all_filters(state['weekly_data'][week], request)

    date_str = request.args.get('date', '')
    try:
        end_date = pd.Timestamp(date_str) if date_str else pd.Timestamp.now().normalize()
    except Exception:
        end_date = pd.Timestamp.now().normalize()

    wq_df = _get_work_queue_df(name)
    exclude_wq = request.args.get('exclude_wq', 'false').lower() == 'true'

    try:
        result = untouched_claims_analysis(atb_df, prod_df, end_date, wq_df=wq_df, exclude_wq=exclude_wq)
    except Exception as e:
        print(f'[workables/download ERROR] {e}')
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    rows = result['rows']
    if not rows:
        return jsonify({'error': 'No unworked claims to download'}), 404

    df = pd.DataFrame(rows)
    mode_label = 'Open' if exclude_wq else 'Unworked'
    mode_title = 'Open ATB Claims (Excl. Work Queue)' if exclude_wq else 'Unworked ATB Claims'
    ctx   = _download_context(name, week, [
        f"{mode_title} — Production window: {result['summary']['prod_window']}",
        f"ATB Week: {week}  |  {mode_label}: {result['summary']['unworked_count']:,} claims  |  Balance: ${result['summary']['unworked_bal']:,.0f}",
    ])
    fname = f"{name}_{mode_label}_ATB_{end_date.strftime('%m%d%Y')}.xlsx".replace(' ', '_')
    return _make_excel_response(df, f'{mode_title} — Workables', ctx, fname)


# Keep old /api/medicare/* paths as aliases so cached bookmarks still work
app.add_url_rule('/api/medicare/trending', view_func=api_trending)
app.add_url_rule('/api/medicare/migration', view_func=api_migration, endpoint='api_migration_alias')
app.add_url_rule('/api/medicare/bifurcation', view_func=api_bifurcation, endpoint='api_bifurcation_alias')


import time as _time

def _friday_auto_refresh():
    """Daemon: every Friday after 8 am, re-discover and reload all client data."""
    last_refresh = None
    _time.sleep(60)  # Let initial load start first
    while True:
        now = datetime.datetime.now()
        today = now.date()
        # weekday() == 4 is Friday
        if today.weekday() == 4 and now.hour >= 8 and today != last_refresh:
            print(f'[FRIDAY-REFRESH] {today} — re-discovering and reloading all client data')
            try:
                discovered = discover_clients()
                for c in discovered:
                    name, folder = c['name'], c['atb_folder']
                    with _clients_lock:
                        existing = _clients.get(name)
                        if existing and existing.get('loading', False):
                            continue
                    t = threading.Thread(target=_reload_client, args=(name, folder), daemon=True)
                    t.start()
                last_refresh = today
                print(f'[FRIDAY-REFRESH] Triggered reload for {len(discovered)} clients')
            except Exception as e:
                print(f'[FRIDAY-REFRESH] Error: {e}')
        _time.sleep(3600)  # Check again in 1 hour


# Start background loading regardless of whether run directly or via gunicorn
threading.Thread(target=_load_all_clients, daemon=True).start()
threading.Thread(target=_friday_auto_refresh, daemon=True, name='friday-refresh').start()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port, use_reloader=False)
